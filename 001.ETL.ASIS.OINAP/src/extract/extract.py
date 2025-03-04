"""
Módulo de Extracción (Extract)
Este módulo se encarga de:
- Descargar archivos Excel desde OneDrive
- Validar la estructura de los archivos descargados
- Extraer los datos crudos de los archivos Excel
"""

import os
import sys
import requests
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
import logging
from datetime import datetime

# Añadir el directorio src al path de Python de forma dinámica
current_dir = os.path.dirname(os.path.abspath(__file__))
src_dir = os.path.dirname(current_dir)
if src_dir not in sys.path:
    sys.path.append(src_dir)

from config_loader import config
import pandas as pd
from openpyxl import load_workbook

class ExcelValidator:
    """
    Clase que provee métodos de validación para asegurar que los archivos Excel
    cumplan con la estructura esperada según la configuración.
    """
    @staticmethod
    def validate_structure(wb, file_path):
        """
        Valida que el archivo Excel cumpla con:
        - Existencia de la pestaña configurada
        - Presencia de la celda de Filial
        - Nombres correctos de las columnas
        """
        try:
            # Validación de la pestaña especificada
            if config.excel['structure']['sheet_name'] not in wb.sheetnames:
                raise ValueError(f"La pestaña '{config.excel['structure']['sheet_name']}' no existe en el archivo")
            
            ws = wb[config.excel['structure']['sheet_name']]
            
            # Validación de la celda que contiene el nombre de la Filial
            filial_cell = config.excel['structure']['filial_cell']
            if not ws[filial_cell].value:
                raise ValueError(f"La celda de Filial ({filial_cell}) está vacía")
            
            # Validación de los nombres de todas las columnas requeridas
            header_row = config.excel['structure']['header_row']
            for field, field_config in config.excel['structure']['columns'].items():
                column = field_config['column']
                expected_name = field_config['expected_name']
                actual_name = ws[f"{column}{header_row}"].value
                
                if actual_name != expected_name:
                    raise ValueError(f"La columna {column} debería llamarse '{expected_name}' pero se encontró '{actual_name}'")
            
            return True
        except Exception as e:
            logging.error(f"Error de validación en {file_path}: {str(e)}")
            return False

class ExcelDownloader:
    """
    Clase principal para la descarga y procesamiento inicial de archivos Excel.
    Maneja la descarga desde OneDrive y el almacenamiento local de los archivos.
    """
    def __init__(self, data_folder):
        """
        Inicializa el descargador con el directorio base para los archivos
        y crea la estructura de carpetas necesaria.
        """
        self.data_folder = data_folder
        self.downloads_folder = None
        self.setup_folders()
    
    def setup_folders(self):
        """
        Crea la estructura de carpetas para almacenar las descargas,
        usando un timestamp único para cada ejecución.
        """
        timestamp = datetime.now().strftime(config.formats['timestamp'])
        self.downloads_folder = os.path.join(self.data_folder, "downloads", f"execution_{timestamp}")
        if not os.path.exists(self.downloads_folder):
            os.makedirs(self.downloads_folder)
    
    def modify_onedrive_link(self, url):
        """
        Modifica las URLs de OneDrive para permitir la descarga directa
        del archivo Excel.
        """
        if '/edit' in url:
            url = url.replace('/edit', '/download')
        
        parsed = urlparse(url)
        query = parse_qs(parsed.query)
        if 'download' not in query:
            query['download'] = ['1']
            new_query = urlencode(query, doseq=True)
            url = urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))
        return url
    
    def download_file(self, url, unique_id):
        """
        Descarga un archivo desde OneDrive y lo guarda localmente.
        Maneja errores de conexión y descarga.
        """
        # Modificar URL si es de OneDrive
        if "onedrive.live.com" in url.lower() or "1drv.ms" in url.lower():
            url = self.modify_onedrive_link(url)
        
        # Generar nombre temporal único para el archivo
        timestamp = datetime.now().strftime(config.formats['timestamp'])
        temp_filename = f"temp_{unique_id}_{timestamp}.xlsx"
        temp_path = os.path.join(self.downloads_folder, temp_filename)

        try:
            # Realizar la descarga en chunks para manejar archivos grandes
            response = requests.get(url, stream=True)
            if response.status_code == 200:
                with open(temp_path, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=config.excel['chunk_size']):
                        if chunk:
                            f.write(chunk)
                logging.info(f"Archivo descargado exitosamente como: {temp_path}")
                return temp_path
            else:
                logging.error(f"Error al descargar {url}: Status {response.status_code}")
                return None
        except Exception as e:
            logging.error(f"Excepción al descargar {url}: {e}")
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass
            return None

    def extract_data(self, file_path, unique_id):
        """
        Extrae y procesa los datos del archivo Excel descargado.
        - Valida la estructura del archivo
        - Renombra el archivo según la filial
        - Lee los datos usando pandas
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            
            # Validar estructura del archivo
            if not ExcelValidator.validate_structure(wb, file_path):
                return None
            
            ws = wb[config.excel['structure']['sheet_name']]
            
            # Obtener nombre de la filial y renombrar archivo
            filial = ws[config.excel['structure']['filial_cell']].value
            if filial:
                filial_name = filial.strip().replace(" ", "_").lower()
                timestamp = datetime.now().strftime(config.formats['timestamp'])
                new_filename = f"archivo_{filial_name}_{unique_id}_{timestamp}.xlsx"
                new_path = os.path.join(self.downloads_folder, new_filename)
                
                os.rename(file_path, new_path)
                file_path = new_path
                logging.info(f"Archivo renombrado para filial {filial}: {new_path}")
            
            # Determinar el rango de datos válidos
            mes_column = config.excel['structure']['columns']['mes']['column']
            header_row = config.excel['structure']['header_row']
            last_row = header_row
            
            for row in range(header_row + 1, ws.max_row + 1):
                cell_value = ws[f"{mes_column}{row}"].value
                if cell_value:
                    last_row = row
                else:
                    break
            
            # Leer datos usando pandas hasta la última fila válida
            df = pd.read_excel(
                file_path, 
                sheet_name=config.excel['structure']['sheet_name'],
                header=config.excel['structure']['header_row']-1,
                nrows=last_row-config.excel['structure']['header_row']+1
            )
            
            return {
                'filial': filial,
                'dataframe': df
            }
            
        except Exception as e:
            logging.error(f"Error al extraer datos de {file_path}: {str(e)}")
            return None
    
    def download_all_files(self, urls):
        """
        Descarga todos los archivos Excel desde las URLs proporcionadas.
        """
        downloaded_files = []
        for idx, url in enumerate(urls, start=1):
            logging.info(f"Iniciando descarga del archivo {idx} desde: {url}")
            file_path = self.download_file(url, idx)
            if file_path:
                downloaded_files.append(file_path)
            else:
                logging.error(f"No se pudo descargar el archivo de: {url}")
        return downloaded_files

    def process_files(self, file_paths):
        """
        Procesa todos los archivos descargados y extrae sus datos.
        """
        extracted_data = []
        for idx, file_path in enumerate(file_paths, 1):
            data = self.extract_data(file_path, idx)
            if data:
                extracted_data.append(data)
            else:
                logging.error(f"No se pudo extraer datos del archivo {file_path}")
        return extracted_data

    def process_urls(self, urls):
        """
        Método principal que coordina todo el proceso de extracción:
        1. Descarga todos los archivos
        2. Procesa cada archivo descargado
        3. Retorna los datos extraídos
        """
        downloaded_files = self.download_all_files(urls)
        logging.info(f"Se descargaron {len(downloaded_files)} archivos exitosamente")
        return self.process_files(downloaded_files)