# etl_excel.py
# Esta aplicación descarga archivos excel de enlaces proporcionados y luego procesa cada archivo para extraer información específica.

import os
import requests
import pandas as pd
from openpyxl import load_workbook, Workbook
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse


def modify_onedrive_link(url):
    """Si el enlace es de OneDrive, modifica el URL para forzar la descarga directa añadiendo el parámetro download=1."""
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    if 'download' not in query:
        query['download'] = ['1']
        new_query = urlencode(query, doseq=True)
        new_url = urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, new_query, parsed.fragment))
        return new_url
    return url


def download_file(url, dest_folder):
    """Descarga un archivo desde la URL y lo guarda en dest_folder."""
    # Si se detecta que el enlace es de OneDrive, modificarlo para descarga directa
    if "onedrive.live.com" in url.lower():
        url = modify_onedrive_link(url)

    if not os.path.exists(dest_folder):
        os.makedirs(dest_folder)
    
    # Extraer nombre de archivo de la URL o asignar uno si no se encuentra
    local_filename = url.split('/')[-1]
    if not local_filename.lower().endswith('.xlsx'):
        local_filename += '.xlsx'
    file_path = os.path.join(dest_folder, local_filename)

    try:
        response = requests.get(url, stream=True)
        # Si el contenido es HTML, es posible que no se haya obtenido el archivo correcto
        if 'text/html' in response.headers.get('Content-Type', ''):
            print(f"El enlace proporcionado devolvió HTML, intentando forzar la descarga directa...")
            url = modify_onedrive_link(url)  # Forzar el parámetro
            response = requests.get(url, stream=True)

        if response.status_code == 200:
            with open(file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            print(f"Descargado: {file_path}")
            return file_path
        else:
            print(f"Error al descargar {url}: Status {response.status_code}")
            return None
    except Exception as e:
        print(f"Excepción al descargar {url}: {e}")
        return None


def extract_cells(file_path):
    """Extrae las celdas H3 hasta H200 del primer sheet de un archivo Excel."""
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        cells = []
        for row in range(3, 201):  # H3 a H200
            cell_value = ws[f'H{row}'].value
            cells.append(cell_value)
        print(f"Extraído datos de {file_path}")
        return cells
    except Exception as e:
        print(f"Error al extraer datos de {file_path}: {e}")
        return []


def update_consolidated(data_columns):
    """Crea o actualiza el fichero consolidated.xlsx agregando una nueva hoja con los datos extraídos."""
    filename = 'consolidated.xlsx'
    if os.path.exists(filename):
        wb = load_workbook(filename)
        new_sheet_name = f'Run {len(wb.sheetnames) + 1}'
    else:
        wb = Workbook()
        # Remover la hoja por defecto
        default_sheet = wb.active
        wb.remove(default_sheet)
        new_sheet_name = 'Run 1'
    
    ws = wb.create_sheet(new_sheet_name)
    
    # Escribir cada columna con su encabezado
    for i, (col_name, data) in enumerate(data_columns.items(), start=1):
        ws.cell(row=1, column=i, value=col_name)
        for j, value in enumerate(data, start=2):
            ws.cell(row=j, column=i, value=value)
    wb.save(filename)
    print(f"Datos consolidados escritos en la hoja '{new_sheet_name}' de {filename}")


def main():
    # Lista de enlaces a archivos Excel almacenados en OneDrive
    urls = [
        "https://1drv.ms/x/s!ArwtqyUha2-Cmnc76ElKT0V_U6VQ?e=9VU2PR"
    ]
    
    local_folder = "downloaded_excels"
    downloaded_files = []
    
    for url in urls:
        file_path = download_file(url, local_folder)
        if file_path:
            downloaded_files.append(file_path)
    
    # Extraer datos y consolidarlos
    data_columns = {}
    for idx, file_path in enumerate(downloaded_files, start=1):
        data = extract_cells(file_path)
        data_columns[f'Archivo {idx}'] = data
    
    update_consolidated(data_columns)


if __name__ == "__main__":
    # Asegurarse de tener instaladas las dependencias: requests, pandas y openpyxl
    # pip install requests pandas openpyxl
    main()
